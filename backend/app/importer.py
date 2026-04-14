"""Importa um arquivo .xlsm no formato Pmais e cria um evento populado.

Uso (dentro do container backend):
    python -m app.importer /files/Pmais.xlsm --name "Radar 2026"
"""
from __future__ import annotations

import argparse
import sys
from decimal import Decimal

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models import Event, ServiceCategory, ServiceItem
from app.seed_template import seed_event_template


def _to_dec(v) -> Decimal | None:
    if v is None or v == "" or v == 0:
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def import_xlsm(path: str, event_name: str | None = None) -> int:
    wb = load_workbook(path, data_only=True, keep_vba=False)
    cfg = wb["Configurações"]
    ctrl = wb["Controle de Evento"]

    cfg_map = {}
    for row in cfg.iter_rows(min_row=1, max_row=20, values_only=True):
        if row[0] and row[1]:
            cfg_map[str(row[0]).strip()] = str(row[1]).strip()

    name = event_name or cfg_map.get("Nome do Evento") or "Evento importado"
    producer_email = cfg_map.get("E-mail do Produtor")
    finance_email = cfg_map.get("E-mail do Financeiro")

    db = SessionLocal()
    try:
        ev = Event(name=name, producer_email=producer_email, finance_email=finance_email)
        db.add(ev); db.flush()
        seed_event_template(db, ev.id)

        # varre linhas de item (coluna A = nome, D..N = valores) e preenche se existir
        cats = {c.name: c for c in db.query(ServiceCategory).filter_by(event_id=ev.id).all()}
        current_cat = None
        for row in ctrl.iter_rows(min_row=9, max_row=ctrl.max_row, values_only=True):
            a = (row[0] or "").strip() if isinstance(row[0], str) else ""
            if not a:
                continue
            if a.isupper() and a not in cats and not a.startswith("TOTAL"):
                # cabeçalho de categoria (ex: "PRÉ EVENTO")
                for k in cats:
                    if k.upper().strip() == a.strip():
                        current_cat = cats[k]; break
                continue
            if a.startswith("TOTAL"):
                continue
            if not current_cat:
                continue
            item = db.query(ServiceItem).filter_by(category_id=current_cat.id, name=a).first()
            if not item:
                continue
            item.planned_qty = _to_dec(row[2])
            item.planned_days = _to_dec(row[3])
            item.planned_unit = _to_dec(row[4])
            item.budgeted_qty = _to_dec(row[6])
            item.budgeted_days = _to_dec(row[7])
            item.budgeted_unit = _to_dec(row[8])
            item.contracted_qty = _to_dec(row[10])
            item.contracted_days = _to_dec(row[11])
            item.contracted_unit = _to_dec(row[12])

        db.commit()
        return ev.id
    finally:
        db.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--name", default=None)
    args = ap.parse_args()
    eid = import_xlsm(args.path, args.name)
    print(f"evento importado id={eid}")

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from jose import JWTError, jwt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session, selectinload

from app.config import settings
from app.database import get_db
from app.models import Event, Revenue, ServiceCategory, User


def _user_from_query(token: str, db: Session) -> User:
    try:
        payload = jwt.decode(token, settings.JWT_SECRET, algorithms=[settings.JWT_ALGORITHM])
        email = payload.get("sub")
    except JWTError:
        raise HTTPException(401, "token inválido")
    u = db.query(User).filter(User.email == email, User.is_active.is_(True)).first()
    if not u:
        raise HTTPException(401, "usuário inválido")
    return u

router = APIRouter(prefix="/api", tags=["reports"])

HDR = PatternFill("solid", fgColor="1D3557")
HDR_FONT = Font(bold=True, color="FFFFFF")
SECTION = PatternFill("solid", fgColor="E2E8F0")


def _total(it, kind: str) -> float:
    q = float(getattr(it, f"{kind}_qty") or 0)
    d = float(getattr(it, f"{kind}_days") or 1)
    u = float(getattr(it, f"{kind}_unit") or 0)
    return q * d * u


@router.get("/events/{event_id}/report.xlsx")
def report_xlsx(event_id: int, token: str, db: Session = Depends(get_db)):
    _user_from_query(token, db)
    ev = db.get(Event, event_id)
    if not ev:
        raise HTTPException(404)
    cats = (
        db.query(ServiceCategory)
        .options(selectinload(ServiceCategory.items))
        .filter(ServiceCategory.event_id == event_id)
        .order_by(ServiceCategory.sort_order)
        .all()
    )
    revs = db.query(Revenue).filter_by(event_id=event_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo Cliente"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"RELATÓRIO — {ev.name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    r = 3
    ws.cell(r, 1, "Evento"); ws.cell(r, 2, ev.name); r += 1
    ws.cell(r, 1, "Data"); ws.cell(r, 2, str(ev.event_date or "")); r += 1
    ws.cell(r, 1, "Local"); ws.cell(r, 2, ev.location or ""); r += 1
    ws.cell(r, 1, "Cliente"); ws.cell(r, 2, ev.client or ""); r += 2

    for col, h in enumerate(["Serviço", "Qtd", "Dias", "Unit", "Total Planej.", "Total Contrat."], start=1):
        c = ws.cell(r, col, h); c.fill = HDR; c.font = HDR_FONT
    r += 1

    total_plan, total_contr = 0.0, 0.0
    for cat in cats:
        ws.cell(r, 1, cat.name).fill = SECTION
        for col in range(2, 7):
            ws.cell(r, col).fill = SECTION
        r += 1
        for it in cat.items:
            tp = _total(it, "planned"); tc = _total(it, "contracted")
            if tp == 0 and tc == 0:
                continue
            ws.cell(r, 1, it.name)
            ws.cell(r, 2, float(it.planned_qty or 0))
            ws.cell(r, 3, float(it.planned_days or 0))
            ws.cell(r, 4, float(it.planned_unit or 0))
            ws.cell(r, 5, tp); ws.cell(r, 6, tc)
            total_plan += tp; total_contr += tc
            r += 1

    r += 1
    ws.cell(r, 1, "TOTAL DESPESAS").font = Font(bold=True)
    ws.cell(r, 5, total_plan).font = Font(bold=True)
    ws.cell(r, 6, total_contr).font = Font(bold=True)
    r += 2

    if revs:
        ws.cell(r, 1, "RECEITAS").font = Font(bold=True, size=12)
        r += 1
        for col, h in enumerate(["Grupo", "Descrição", "Qtd Plan.", "Unit", "Total Plan.", "Total Real."], start=1):
            c = ws.cell(r, col, h); c.fill = HDR; c.font = HDR_FONT
        r += 1
        trp, trr = 0.0, 0.0
        for rv in revs:
            tp = float(rv.planned_qty or 0) * float(rv.planned_unit or 0)
            tr = float(rv.realized_qty or 0) * float(rv.realized_unit or 0)
            ws.cell(r, 1, rv.group_name)
            ws.cell(r, 2, rv.name)
            ws.cell(r, 3, float(rv.planned_qty or 0))
            ws.cell(r, 4, float(rv.planned_unit or 0))
            ws.cell(r, 5, tp); ws.cell(r, 6, tr)
            trp += tp; trr += tr
            r += 1
        ws.cell(r, 1, "TOTAL RECEITAS").font = Font(bold=True)
        ws.cell(r, 5, trp).font = Font(bold=True)
        ws.cell(r, 6, trr).font = Font(bold=True)
        r += 2
        ws.cell(r, 1, "SALDO PLANEJADO").font = Font(bold=True)
        ws.cell(r, 5, trp - total_plan).font = Font(bold=True)

    for col_letter, width in zip("ABCDEF", [40, 12, 10, 14, 18, 18]):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"relatorio-{ev.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
